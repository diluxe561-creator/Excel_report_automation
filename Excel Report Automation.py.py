import openpyxl

# -----------------------------
# STEP 1: Create sales_data.xlsx
# -----------------------------
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Sales Data"

# Add headers
sheet["A1"] = "Date"
sheet["B1"] = "Sales"
sheet["A5"] = "Total sales"
sheet["B5"] = 3400

# Add sample sales data
data = [
    ("2025-01-01", 1200,),
    ("2025-01-02", 850),
    ("2025-01-03", 640),
    ("2025-01-04", 910),
]

for row in data:
    sheet.append(row)

wb.save("sales_data.xlsx")

# -----------------------------
# STEP 2: Read sales_data.xlsx
# -----------------------------
wb = openpyxl.load_workbook("sales_data.xlsx")
sheet = wb.active

total_sales = 0

for row in range(2, sheet.max_row + 1):
    value = sheet.cell(row=row, column=2).value
    if value is not None:
        total_sales += value

# -----------------------------
# STEP 3: Create sales_summary.xlsx
# -----------------------------
summary_wb = openpyxl.Workbook()
summary_sheet = summary_wb.active

summary_sheet["A1"] = "Total Sales"
summary_sheet["B1"] = total_sales

summary_wb.save("sales_summary.xlsx")

print("Sales report created successfully!")
